from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── SHEET 1: Project Table ────────────────────────────────────────────────────
ws = wb.active
ws.title = "CapEx Tracker"

# Colour palette
CLR_HEADER      = "1F3864"   # dark navy
CLR_HEADER_FONT = "FFFFFF"
CLR_AT_RISK     = "FF4C4C"   # red
CLR_WATCH       = "FFB347"   # orange
CLR_NOT_STARTED = "D3D3D3"   # light grey
CLR_FLAGGED     = "FFD700"   # gold
CLR_ALT_ROW     = "EEF2F8"   # light blue-grey
CLR_TITLE_BG    = "2E74B5"
CLR_SUMMARY_HDR = "D6E4F7"

thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="1F3864")

def border(left=thin, right=thin, top=thin, bottom=thin):
    return Border(left=left, right=right, top=top, bottom=bottom)

def hdr_font(sz=11):
    return Font(name="Calibri", bold=True, color=CLR_HEADER_FONT, size=sz)

def body_font(bold=False, sz=10):
    return Font(name="Calibri", bold=bold, size=sz)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Title row ────────────────────────────────────────────────────────────────
ws.merge_cells("A1:O1")
title = ws["A1"]
title.value = "CapEx Project Tracker — Multifamily Portfolio"
title.font  = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
title.fill  = fill(CLR_TITLE_BG)
title.alignment = center()
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:O2")
sub = ws["A2"]
sub.value = "Last Updated: March 25, 2026  |  Total Budget: $720,000  |  Spent: $243,500  |  Remaining: $476,500  |  % Spent: 33.8%"
sub.font  = Font(name="Calibri", bold=False, size=10, color="1F3864")
sub.fill  = fill("D6E4F7")
sub.alignment = center()
ws.row_dimensions[2].height = 18

# ── Column headers ───────────────────────────────────────────────────────────
headers = [
    "Project Name", "Property", "Scope", "Total Budget", "Spent to Date",
    "Budget Remaining", "Vendor / Contractor", "Start Date", "Est. Completion",
    "Owner's Rep", "Status", "% Complete", "Priority", "Next Action", "Notes"
]
col_widths = [28, 16, 22, 14, 14, 16, 22, 12, 16, 14, 14, 12, 10, 40, 36]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font      = hdr_font()
    cell.fill      = fill(CLR_HEADER)
    cell.alignment = center()
    cell.border    = border(left=Side(style="medium", color="FFFFFF"),
                            right=Side(style="medium", color="FFFFFF"),
                            top=thin, bottom=thin)
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[3].height = 30

# ── Project data ─────────────────────────────────────────────────────────────
# (name, property, scope, budget, spent, vendor, start, est_comp, owner, status, pct, priority, next_action, notes)
projects = [
    ("Park Layne Structural",       "Park Layne",     "Structural",              350000, 230000, "Hart Restoration",   "—", "Apr 15, 2026", "Nate Fisher",  "AT RISK",     0.65, "HIGH",   "Get written milestone schedule from Hart confirming $120K scope closes by Apr 15",                                      "21 days to deadline with $120K remaining — aggressive"),
    ("Park Layne Insurance",        "Park Layne",     "Insurance Work",          100000,      0, "Sturgeon Contracting","—", "Apr 15, 2026", "Stephen B",    "AT RISK",     0.00, "HIGH",   "Call Sturgeon today — confirm start date in writing",                                                                  "$0 spent against Apr 15 deadline is a critical gap"),
    ("Park Layne Roof",             "Park Layne",     "Roofing",                 100000,   5500, "NONE",               "—", "May 1, 2026",  "Stephen B",    "AT RISK",     0.05, "HIGH",   "Solicit 3 roofing bids immediately; award by Apr 1",                                                                   "No contractor — $94.5K of work with 37 days to go"),
    ("Executive House 808 Fire Unit","Executive House","Fire / Interior Restoration",60000, 3000, "NONE",              "—", "May 1, 2026",  "UNASSIGNED",   "AT RISK",     0.05, "HIGH",   "Assign Owner's Rep today; hire GC this week; initiate MEP coordination",                                               "No OR, no GC — fire/life safety scope cannot sit unmanaged"),
    ("Executive House 803",         "Executive House","Renovation",               15000,   4000, "NONE",               "—", "NONE",         "Stephen B",    "AT RISK",     0.25, "MEDIUM", "Hire GC and set hard completion date — plans submitted, no excuse for further delay",                                   "No contractor, no timeline"),
    ("Riverstone 309",              "Riverstone",     "Interior — Kitchen & Bath",15000,      0, "NONE",               "—", "Apr 15, 2026", "Stephen B",    "AT RISK",     0.00, "HIGH",   "Award interior contractor this week — consider same contractor as 403 for concurrent runs",                            "$0 spent, no contractor, 21 days out"),
    ("Riverstone 403",              "Riverstone",     "Interior — Kitchen & Bath",15000,      0, "NONE",               "—", "Apr 15, 2026", "Stephen B",    "AT RISK",     0.00, "HIGH",   "Hire interior contractor immediately — same urgency as 309; confirm one contractor can run both units concurrently",   "Separate unit confirmed; same scope as 309"),
    ("Riverstone Steps",            "Riverstone",     "Exterior — Concrete Steps",15000,   1000, "NONE",               "—", "May 1, 2026",  "Stephen B",    "AT RISK",     0.07, "MEDIUM", "Hire concrete contractor; $1K likely mobilization/demo only",                                                          "No contractor assigned"),
    ("Executive House Fire Panel",  "Executive House","Fire Panel System",        50000,      0, "Koorsen Fire",       "—", "May 1, 2026",  "Nate Fisher",  "NOT STARTED", 0.00, "HIGH",   "Confirm Koorsen mobilization date and permit status — fire panel is life safety, no slippage acceptable",             "Vendor assigned; must confirm start"),
]

STATUS_COLORS = {
    "AT RISK":     "FF4C4C",
    "NOT STARTED": "D3D3D3",
    "IN PROGRESS": "70AD47",
    "COMPLETE":    "375623",
}
PRIORITY_COLORS = {
    "HIGH":   "FF4C4C",
    "MEDIUM": "FFB347",
    "LOW":    "70AD47",
}

for row_idx, p in enumerate(projects, start=4):
    name, prop, scope, budget, spent, vendor, start, comp, owner, status, pct, priority, next_action, notes = p
    remaining = budget - spent
    row_fill = fill(CLR_ALT_ROW) if row_idx % 2 == 0 else fill("FFFFFF")

    data = [name, prop, scope, budget, spent, remaining, vendor, start, comp, owner, status, pct, priority, next_action, notes]

    for col_idx, val in enumerate(data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font      = body_font()
        cell.alignment = center() if col_idx in (1,2,3,4,5,6,8,9,10,11,12,13) else left()
        cell.border    = border()
        cell.fill      = row_fill

        # Currency columns
        if col_idx in (4, 5, 6):
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal="right", vertical="center")

        # % complete
        if col_idx == 12:
            cell.number_format = '0%'

        # Status colour
        if col_idx == 11:
            hex_c = STATUS_COLORS.get(status, "FFFFFF")
            cell.fill = fill(hex_c)
            cell.font = Font(name="Calibri", bold=True, size=10,
                             color="FFFFFF" if status in ("AT RISK",) else "000000")

        # Priority colour
        if col_idx == 13:
            hex_c = PRIORITY_COLORS.get(priority, "FFFFFF")
            cell.fill = fill(hex_c)
            cell.font = Font(name="Calibri", bold=True, size=10,
                             color="FFFFFF" if priority == "HIGH" else "000000")

        # Flag missing vendor / owner's rep
        if col_idx in (7, 10) and val in ("NONE", "UNASSIGNED"):
            cell.fill = fill("FF4C4C")
            cell.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")

    ws.row_dimensions[row_idx].height = 42

# Freeze panes below header
ws.freeze_panes = "A4"


# ── SHEET 2: Risk & Actions ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Risk & Actions")

ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "Risk Summary & Next Actions — As of March 25, 2026"
t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
t.fill  = fill(CLR_TITLE_BG)
t.alignment = center()
ws2.row_dimensions[1].height = 28

# Section: At-Risk Projects
sections = [
    ("🚨 AT-RISK PROJECTS", "FF4C4C", [
        ("Park Layne Insurance",         "Stephen B",  "CRITICAL", "Call Sturgeon Contracting today — confirm start date in writing. $0 spent, Apr 15 deadline, 21 days out."),
        ("Executive House 808 Fire Unit","UNASSIGNED", "CRITICAL", "Assign Owner's Rep today. Hire GC this week. Initiate MEP coordination. Fire/life safety — cannot sit unmanaged."),
        ("Riverstone 309",               "Stephen B",  "CRITICAL", "Award interior contractor this week. Consider same contractor as 403 to run both units concurrently."),
        ("Riverstone 403",               "Stephen B",  "CRITICAL", "Hire interior contractor immediately. Same urgency as 309. Apr 15 is 21 days out with $0 spent."),
        ("Park Layne Structural",        "Nate Fisher","HIGH",     "Get written milestone schedule from Hart confirming $120K of work closes by Apr 15."),
        ("Park Layne Roof",              "Stephen B",  "HIGH",     "Solicit 3 roofing bids immediately. Award contract by Apr 1 or May 1 deadline is lost."),
        ("Executive House 803",          "Stephen B",  "HIGH",     "Hire GC and set hard completion date. Plans are submitted — no reason for further delay."),
        ("Riverstone Steps",             "Stephen B",  "MEDIUM",   "Hire concrete contractor. $1K of spend is likely mobilization/demo only."),
    ]),
    ("💰 BUDGET WATCH", "2E74B5", [
        ("Park Layne Structural",        "$230K / $350K", "65.7% spent", "$120K remaining with 21 days to go — confirm scope is achievable."),
        ("Park Layne Insurance",         "$0 / $100K",    "0% spent",    "Full budget uncommitted, deadline in 21 days."),
        ("Park Layne Roof",              "$5.5K / $100K", "5.5% spent",  "$94.5K uncommitted, no contractor."),
        ("Executive House 808",          "$3K / $60K",    "5% spent",    "Fire scope, no GC — $57K at risk."),
        ("Executive House Fire Panel",   "$0 / $50K",     "0% spent",    "Vendor assigned; must confirm mobilization."),
        ("PORTFOLIO TOTAL",              "$243.5K / $720K","33.8% spent","$476.5K uncommitted across 9 projects."),
    ]),
    ("⏱ TIMELINE ISSUES", "7030A0", [
        ("Park Layne Insurance",         "Apr 15",  "21 days", "0% spent — CRITICAL"),
        ("Riverstone 309",               "Apr 15",  "21 days", "0% spent, no contractor — CRITICAL"),
        ("Riverstone 403",               "Apr 15",  "21 days", "0% spent, no contractor — CRITICAL"),
        ("Park Layne Structural",        "Apr 15",  "21 days", "$120K remaining — HIGH"),
        ("Park Layne Roof",              "May 1",   "37 days", "No contractor — HIGH"),
        ("Executive House 808",          "May 1",   "37 days", "No GC, no OR — HIGH"),
        ("Executive House Fire Panel",   "May 1",   "37 days", "Vendor assigned, not started — MEDIUM"),
        ("Riverstone Steps",             "May 1",   "37 days", "No contractor — MEDIUM"),
        ("Executive House 803",          "NONE",    "—",       "No completion date set — HIGH"),
    ]),
]

current_row = 3
for section_title, color, rows in sections:
    # Section header
    ws2.merge_cells(f"A{current_row}:F{current_row}")
    sh = ws2[f"A{current_row}"]
    sh.value = section_title
    sh.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    sh.fill  = fill(color)
    sh.alignment = left()
    ws2.row_dimensions[current_row].height = 22
    current_row += 1

    # Column sub-headers
    for col_idx, hdr in enumerate(["Project", "Owner / Spent", "Status / Days", "Action / Note"], start=1):
        c = ws2.cell(row=current_row, column=col_idx, value=hdr)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = fill(CLR_HEADER)
        c.alignment = center()
        c.border    = border()
    ws2.row_dimensions[current_row].height = 18
    current_row += 1

    for i, row_data in enumerate(rows):
        row_bg = fill(CLR_ALT_ROW) if i % 2 == 0 else fill("FFFFFF")
        for col_idx, val in enumerate(row_data, start=1):
            c = ws2.cell(row=current_row, column=col_idx, value=val)
            c.font      = body_font()
            c.fill      = row_bg
            c.alignment = left()
            c.border    = border()
            if col_idx == 3 and "CRITICAL" in str(val):
                c.fill = fill("FF4C4C")
                c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        ws2.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 1  # spacer

# Column widths for sheet 2
for col_letter, w in zip(["A","B","C","D"], [30, 20, 18, 55]):
    ws2.column_dimensions[col_letter].width = w

ws2.freeze_panes = "A3"


# ── SHEET 3: Next Actions by Owner ───────────────────────────────────────────
ws3 = wb.create_sheet("Next Actions")

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "Next Actions — Grouped by Owner's Rep"
t3.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
t3.fill  = fill(CLR_TITLE_BG)
t3.alignment = center()
ws3.row_dimensions[1].height = 28

owners = [
    ("Nate Fisher", "1F3864", [
        ("Park Layne Structural",      "HIGH",   "Get written milestone schedule from Hart confirming $120K scope closes by Apr 15. Immediately."),
        ("Executive House Fire Panel", "HIGH",   "Confirm Koorsen mobilization date and permit status. Fire panel is life safety — no slippage acceptable."),
    ]),
    ("Stephen B", "2E74B5", [
        ("Park Layne Insurance",  "HIGH",   "Call Sturgeon Contracting TODAY — confirm start date in writing. Apr 15 is 21 days out with $0 spent."),
        ("Park Layne Roof",       "HIGH",   "Solicit 3 roofing bids immediately. Award contract by Apr 1 hard deadline."),
        ("Executive House 803",   "MEDIUM", "Hire GC and set hard completion date this week. Plans are submitted."),
        ("Riverstone 309",        "HIGH",   "Award interior contractor this week. Consider same contractor as 403 for concurrent execution."),
        ("Riverstone 403",        "HIGH",   "Hire interior contractor immediately. Same urgency as 309."),
        ("Riverstone Steps",      "MEDIUM", "Hire concrete contractor. Provide mobilization date."),
    ]),
    ("UNASSIGNED — Escalate Now", "FF4C4C", [
        ("Executive House 808 Fire Unit", "HIGH", "Assign Owner's Rep TODAY. Then: hire GC, initiate MEP coordination. Fire/life safety — cannot be managed by committee."),
    ]),
]

current_row = 3
for owner_name, color, actions in owners:
    ws3.merge_cells(f"A{current_row}:D{current_row}")
    oh = ws3[f"A{current_row}"]
    oh.value = owner_name
    oh.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    oh.fill  = fill(color)
    oh.alignment = left()
    ws3.row_dimensions[current_row].height = 22
    current_row += 1

    for col_idx, hdr in enumerate(["Project", "Priority", "Action"], start=1):
        c = ws3.cell(row=current_row, column=col_idx, value=hdr)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = fill(CLR_HEADER)
        c.alignment = center()
        c.border    = border()
    ws3.row_dimensions[current_row].height = 18
    current_row += 1

    for i, (proj, pri, action) in enumerate(actions):
        row_bg = fill(CLR_ALT_ROW) if i % 2 == 0 else fill("FFFFFF")
        for col_idx, val in enumerate([proj, pri, action], start=1):
            c = ws3.cell(row=current_row, column=col_idx, value=val)
            c.font      = body_font()
            c.fill      = row_bg
            c.alignment = left()
            c.border    = border()
            if col_idx == 2:
                hex_c = PRIORITY_COLORS.get(pri, "FFFFFF")
                c.fill = fill(hex_c)
                c.font = Font(name="Calibri", bold=True, size=10,
                              color="FFFFFF" if pri == "HIGH" else "000000")
                c.alignment = center()
        ws3.row_dimensions[current_row].height = 30
        current_row += 1

    current_row += 1

for col_letter, w in zip(["A","B","C"], [30, 10, 70]):
    ws3.column_dimensions[col_letter].width = w

ws3.freeze_panes = "A3"

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = "/home/user/clawbot/capex_tracker.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
