#!/usr/bin/env python3
"""Generate Roxster, Ltd. Financial Analysis Excel workbook."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()

# ── Style constants ──────────────────────────────────────────────────────────
NAVY = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True, size=11)
BOLD_TOTAL = Font(bold=True, size=11, color="1F3864")
NORMAL = Font(size=11)
CURRENCY_FMT = '$#,##0.00'
CURRENCY_INT = '$#,##0'
PCT_FMT = '0.0%'
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="AAAAAA"),
)
THICK_BORDER = Border(
    top=Side(style="medium", color="1F3864"),
    bottom=Side(style="medium", color="1F3864"),
)

def header_row(ws, row, cols, texts):
    for c, t in zip(cols, texts):
        cell = ws.cell(row=row, column=c, value=t)
        cell.fill = NAVY
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

def curr(ws, row, col, val, fmt=CURRENCY_FMT, font=None, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = fmt
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell

def pct(ws, row, col, val, font=None, fill=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.number_format = PCT_FMT
    if font: cell.font = font
    if fill: cell.fill = fill
    return cell

def label(ws, row, col, val, font=NORMAL, indent=0, fill=None, border=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = font
    cell.alignment = Alignment(indent=indent)
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def shade_alt(ws, row, max_col):
    if row % 2 == 0:
        for c in range(1, max_col + 1):
            if ws.cell(row=row, column=c).fill == PatternFill():
                ws.cell(row=row, column=c).fill = LIGHT_GRAY


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Income Statement
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Income Statement"
set_col_widths(ws1, [34, 18, 18, 16, 12, 18, 16, 12])

# Title block
ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value="F & W Properties Inc.").font = Font(bold=True, size=14, color="1F3864")
ws1.cell(row=1, column=1).alignment = Alignment(horizontal="center")
ws1.merge_cells("A2:H2")
ws1.cell(row=2, column=1, value="Monthly Income Statement — March 2026").font = Font(bold=True, size=12, color="1F3864")
ws1.cell(row=2, column=1).alignment = Alignment(horizontal="center")
ws1.merge_cells("A3:H3")
ws1.cell(row=3, column=1, value="Roxster, Ltd. (Roxbury Arms)").font = Font(bold=True, size=12, color="1F3864")
ws1.cell(row=3, column=1).alignment = Alignment(horizontal="center")

r = 5
header_row(ws1, r, range(1, 9), [
    "Account", "Mar 2026 Budget", "Mar 2025 Actual",
    "Bgt vs PY ($)", "Bgt vs PY (%)", "Mar 2024 Actual",
    "YoY ($)", "YoY (%)"
])
ws1.freeze_panes = "A6"

def income_line(ws, row, name, budget, act25=None, act24=None, indent=1, bold=False,
                border=None, is_total=False):
    f = BOLD_TOTAL if is_total else (BOLD if bold else NORMAL)
    label(ws, row, 1, name, font=f, indent=indent, border=border)
    curr(ws, row, 2, budget, font=f, border=border)
    if act25 is not None:
        curr(ws, row, 3, act25, font=f, border=border)
        diff = budget - act25
        curr(ws, row, 4, diff, font=f, border=border)
        pct(ws, row, 5, diff / act25 if act25 != 0 else 0, font=f)
    if act24 is not None:
        curr(ws, row, 6, act24, font=f, border=border)
        if act25 is not None:
            yoy = act25 - act24
            curr(ws, row, 7, yoy, font=f, border=border)
            pct(ws, row, 8, yoy / act24 if act24 != 0 else 0, font=f)

r = 6
label(ws1, r, 1, "OPERATING INCOME", font=BOLD); r += 1

# --- Rents detail ---
income_line(ws1, r, "Gross Potential Rent", 121824); r += 1
income_line(ws1, r, "Loss to Vacancy (5%)", -6091.20); r += 1
income_line(ws1, r, "HUD Rent", 885); r += 1
income_line(ws1, r, "Loss to Lease (10%)", -12182.40); r += 1
income_line(ws1, r, "Pet Rent", 864); r += 1
income_line(ws1, r, "Concession", -146); r += 1
income_line(ws1, r, "Total RENTS", 105153.40, 98392.77, 99998.70,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

# --- Fee Income ---
income_line(ws1, r, "NSF Fees", 22); r += 1
income_line(ws1, r, "Application Fee Income", 250); r += 1
income_line(ws1, r, "Tenant Liability Insurance", 272); r += 1
income_line(ws1, r, "Late Fees", 319); r += 1
income_line(ws1, r, "Total FEE INCOME", 863, 918.00, 501.12,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

# --- Other Income ---
income_line(ws1, r, "LLI Admin", 86, 100.50, 42.00); r += 2

# --- Utility Reimbursement ---
income_line(ws1, r, "Water/Sewer Income", 3385); r += 1
income_line(ws1, r, "Electric Income", 7); r += 1
income_line(ws1, r, "Gas & Trash Income", 58); r += 1
income_line(ws1, r, "Total UTILITY REIMBURSEMENT", 3450, 3423.44, 3602.12,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

# --- Total Operating Income ---
income_line(ws1, r, "TOTAL OPERATING INCOME", 110238.40, 102834.71, 104143.94,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

# ── OPERATING EXPENSES ──
label(ws1, r, 1, "OPERATING EXPENSES", font=BOLD); r += 1

income_line(ws1, r, "Marketing", 62, 300.00, 365.78); r += 1
income_line(ws1, r, "Phone & Voicemail", 49, 47.78, 47.57); r += 1
income_line(ws1, r, "Tenant Screening", 7, 0, 0); r += 1
income_line(ws1, r, "Total Administrative", 118, 347.78, 413.35,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "Leasing & Support", 14394, 0, 15095.74); r += 1
income_line(ws1, r, "Grounds Labor", 624, 646.02, 1783.35); r += 1
income_line(ws1, r, "Punchout Labor", 5168, 0, 8767.59); r += 1
income_line(ws1, r, "Maintenance Labor", 2826, 5444.89, 3311.77); r += 1
income_line(ws1, r, "Total Payroll", 23012, 6090.91, 28958.45,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "Electricity", 1177, 1179.10, 1231.62); r += 1
income_line(ws1, r, "Gas", 1360, 2341.69, 1387.27); r += 1
income_line(ws1, r, "Water & Sewer", 2152, 2160.13, 6902.80); r += 1
income_line(ws1, r, "Garbage and Recycling", 1485, 1431.37, 973.29); r += 1
income_line(ws1, r, "Total Utilities", 6174, 7112.29, 10494.98,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "Total R&M", 14043, 22182.18, 12036.72,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

# Taxes & Insurance detail
income_line(ws1, r, "Property Tax", 3454, 3353.43, 3300.37); r += 1
income_line(ws1, r, "Property Insurance", 3658, 1439.98, 1635.83); r += 1
income_line(ws1, r, "Commercial Activity Tax", 66); r += 1
income_line(ws1, r, "City & Local Taxes", 1271); r += 1
income_line(ws1, r, "Total Taxes & Insurance", 8449, 4793.41, 4936.20,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "Legal", 30); r += 1
income_line(ws1, r, "Other Professional Fees", 783); r += 1
income_line(ws1, r, "Bank & Merchant Fees", 4); r += 1
income_line(ws1, r, "Management Fees", 5477.62); r += 1
income_line(ws1, r, "Total Professional Services", 6294.62, 6943.50, 5136.94,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "TOTAL OPERATING EXPENSE", 56753.62, 47470.07, 61976.64,
            indent=0, is_total=True, border=THICK_BORDER); r += 2

income_line(ws1, r, "NET OPERATING INCOME", 53484.78, 55364.64, 42167.30,
            indent=0, is_total=True, border=THICK_BORDER)

# Alt row shading for Sheet1
for row_idx in range(6, r + 1):
    shade_alt(ws1, row_idx, 8)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Valuation
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Valuation")
set_col_widths(ws2, [28, 18, 16, 16, 16, 16, 16, 16])

ws2.merge_cells("A1:H1")
ws2.cell(row=1, column=1, value="Roxster, Ltd. — Valuation Analysis").font = Font(bold=True, size=14, color="1F3864")
ws2.cell(row=1, column=1).alignment = Alignment(horizontal="center")

r = 3
label(ws2, r, 1, "KEY ASSUMPTIONS", font=BOLD); r += 1
assumptions = [
    ("Adjusted Annual NOI", 628654.12),
    ("Cap Rate", 0.075),
    ("Estimated Property Value", 8382054.93),
    ("Outstanding Debt", 5200000.00),
    ("Closing Costs (3%)", 251461.65),
    ("Est. Equity Residual", 2930593.29),
]
for name, val in assumptions:
    label(ws2, r, 1, name, font=NORMAL, indent=1)
    cell = ws2.cell(row=r, column=2, value=val)
    if name == "Cap Rate":
        cell.number_format = PCT_FMT
    else:
        cell.number_format = CURRENCY_FMT
    cell.font = BOLD if "Equity" in name else NORMAL
    r += 1

r += 2
label(ws2, r, 1, "SENSITIVITY TABLE — Equity Residual by Cap Rate", font=BOLD)
r += 1

cap_rates = [0.06, 0.065, 0.07, 0.075, 0.08, 0.085]
scenarios = [
    ("Stress (−10%)", 565788.71),
    ("Base (−5%)", 597221.41),
    ("Budget", 628654.12),
    ("Upside (+5%)", 660086.83),
    ("Upside (+10%)", 691519.53),
]

# Header
label(ws2, r, 1, "NOI Scenario", font=WHITE_FONT)
ws2.cell(row=r, column=1).fill = NAVY
label(ws2, r, 2, "NOI", font=WHITE_FONT)
ws2.cell(row=r, column=2).fill = NAVY
for i, cr in enumerate(cap_rates):
    cell = ws2.cell(row=r, column=3 + i, value=cr)
    cell.number_format = '0.0%'
    cell.fill = NAVY
    cell.font = WHITE_FONT
    cell.alignment = Alignment(horizontal="center")
r += 1

for scenario_name, noi in scenarios:
    label(ws2, r, 1, scenario_name, font=BOLD if scenario_name == "Budget" else NORMAL)
    curr(ws2, r, 2, noi, fmt=CURRENCY_INT)
    for i, cr in enumerate(cap_rates):
        equity = (noi / cr) * (1 - 0.03) - 5200000
        cell = curr(ws2, r, 3 + i, equity, fmt=CURRENCY_INT)
        # Color coding
        if scenario_name == "Budget" and cr == 0.075:
            cell.fill = GOLD_FILL
        elif equity > 3000000:
            cell.fill = GREEN_FILL
        elif equity < 1500000:
            cell.fill = RED_FILL
    r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: 12-Month Budget
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("12-Month Budget")
set_col_widths(ws3, [30] + [16] * 13)

ws3.merge_cells("A1:N1")
ws3.cell(row=1, column=1, value="Roxster, Ltd. — 12-Month Operating Budget (2026)").font = Font(bold=True, size=14, color="1F3864")
ws3.cell(row=1, column=1).alignment = Alignment(horizontal="center")

months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Total"]

r = 3
header_row(ws3, r, range(1, 15), ["Account"] + months)
ws3.freeze_panes = "B4"
r = 4

def budget_row(ws, row, name, h1_val, h2_val, font=NORMAL, indent=1, border=None):
    """Write a monthly row with h1 value for Jan-Jun, h2 for Jul-Dec, and total."""
    label(ws, row, 1, name, font=font, indent=indent, border=border)
    total = 0
    for m in range(1, 13):
        val = h1_val if m <= 6 else h2_val
        total += val
        curr(ws, row, 1 + m, val, font=font, border=border)
    curr(ws, row, 14, total, font=font, border=border)
    return total

# ── INCOME ──
label(ws3, r, 1, "OPERATING INCOME", font=BOLD); r += 1
budget_row(ws3, r, "Total RENTS", 105153.40, 107224.41); r += 1
budget_row(ws3, r, "Total FEE INCOME", 863, 863); r += 1
budget_row(ws3, r, "Other Income (LLI Admin)", 86, 86); r += 1  # using 86 as per data "Other Income: $772/mo" — wait, let me check: LLI Admin $86 but "Other Income: $772/mo"
# Actually the data says Other Income: $772/mo. Let me fix:
# Overwrite the Other Income row
r -= 1  # go back
budget_row(ws3, r, "Other Income", 772, 772); r += 1
budget_row(ws3, r, "Total UTILITY REIMBURSEMENT", 3450, 3450); r += 1

# Total Operating Income
label(ws3, r, 1, "Total Operating Income", font=BOLD_TOTAL, border=THICK_BORDER)
for m in range(1, 13):
    val = 110238.40 if m <= 6 else 112309.41
    curr(ws3, r, 1 + m, val, font=BOLD_TOTAL, border=THICK_BORDER)
curr(ws3, r, 14, 110238.40 * 6 + 112309.41 * 6, font=BOLD_TOTAL, border=THICK_BORDER)
r += 2

# ── EXPENSES ──
label(ws3, r, 1, "OPERATING EXPENSES", font=BOLD); r += 1
budget_row(ws3, r, "Administrative", 118, 118); r += 1
budget_row(ws3, r, "Payroll", 23012, 23012); r += 1
budget_row(ws3, r, "Utilities", 6174, 6174); r += 1
budget_row(ws3, r, "R&M", 14043, 14043); r += 1
budget_row(ws3, r, "Taxes & Insurance", 8449, 8449); r += 1
# Show T&I detail sub-rows
for sub_name, sub_val in [("Property Tax", 3454), ("Property Insurance", 3658),
                           ("Commercial Activity Tax", 66), ("City & Local Taxes", 1271)]:
    label(ws3, r, 1, sub_name, font=Font(size=10, italic=True, color="666666"), indent=2)
    for m in range(1, 13):
        c = ws3.cell(row=r, column=1 + m, value=sub_val)
        c.number_format = CURRENCY_INT
        c.font = Font(size=10, italic=True, color="666666")
    c = ws3.cell(row=r, column=14, value=sub_val * 12)
    c.number_format = CURRENCY_INT
    c.font = Font(size=10, italic=True, color="666666")
    r += 1

budget_row(ws3, r, "Professional Services", 6294.62, 6398.17); r += 1

# Total Operating Expense
label(ws3, r, 1, "Total Operating Expense", font=BOLD_TOTAL, border=THICK_BORDER)
for m in range(1, 13):
    val = 56753.62 if m <= 6 else 56857.17
    curr(ws3, r, 1 + m, val, font=BOLD_TOTAL, border=THICK_BORDER)
curr(ws3, r, 14, 56753.62 * 6 + 56857.17 * 6, font=BOLD_TOTAL, border=THICK_BORDER)
r += 2

# ── NOI ──
label(ws3, r, 1, "NET OPERATING INCOME", font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)
for m in range(1, 13):
    inc = 110238.40 if m <= 6 else 112309.41
    exp = 56753.62 if m <= 6 else 56857.17
    curr(ws3, r, 1 + m, inc - exp, font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)
annual_noi = (110238.40 - 56753.62) * 6 + (112309.41 - 56857.17) * 6
curr(ws3, r, 14, annual_noi, font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)
r += 2

# ── Below the Line ──
label(ws3, r, 1, "BELOW THE LINE", font=BOLD); r += 1

# Mortgage Interest — linear interpolation from $18,315.51 (Jan) to $18,216.78 (Dec)
int_jan = 18315.51
int_dec = 18216.78
int_total_target = 219198.56
# Create a linear schedule
int_monthly = [int_jan + (int_dec - int_jan) * i / 11 for i in range(12)]
# Scale to match total
int_sum = sum(int_monthly)
scale = int_total_target / int_sum
int_monthly = [v * scale for v in int_monthly]

label(ws3, r, 1, "Mortgage Interest", font=NORMAL, indent=1)
for m in range(12):
    curr(ws3, r, 2 + m, int_monthly[m])
curr(ws3, r, 14, sum(int_monthly))
r += 1

# Principal Reduction
prin_jan = 1906.71
prin_dec = 2005.44
prin_total_target = 23468.08
prin_monthly = [prin_jan + (prin_dec - prin_jan) * i / 11 for i in range(12)]
prin_sum = sum(prin_monthly)
scale_p = prin_total_target / prin_sum
prin_monthly = [v * scale_p for v in prin_monthly]

label(ws3, r, 1, "Principal Reduction", font=NORMAL, indent=1)
for m in range(12):
    curr(ws3, r, 2 + m, prin_monthly[m])
curr(ws3, r, 14, sum(prin_monthly))
r += 1

# CapEx
capex = [75000, 75000, 20000, 20000, 20000, 95000, 20000, 10000, 6000, 0, 0, 0]
label(ws3, r, 1, "Capital Expenditures", font=NORMAL, indent=1)
for m in range(12):
    curr(ws3, r, 2 + m, capex[m], fmt=CURRENCY_INT)
curr(ws3, r, 14, sum(capex), fmt=CURRENCY_INT)
r += 1

# Total Below the Line
label(ws3, r, 1, "Total Below the Line", font=BOLD_TOTAL, border=THICK_BORDER)
for m in range(12):
    total_btl = int_monthly[m] + prin_monthly[m] + capex[m]
    curr(ws3, r, 2 + m, total_btl, font=BOLD_TOTAL, border=THICK_BORDER)
curr(ws3, r, 14, sum(int_monthly) + sum(prin_monthly) + sum(capex),
     font=BOLD_TOTAL, border=THICK_BORDER)
r += 2

# Net Cash Flow
label(ws3, r, 1, "NET CASH FLOW", font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)
for m in range(12):
    inc = 110238.40 if m <= 5 else 112309.41
    exp = 56753.62 if m <= 5 else 56857.17
    btl = int_monthly[m] + prin_monthly[m] + capex[m]
    ncf = inc - exp - btl
    curr(ws3, r, 2 + m, ncf, font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)
annual_ncf = annual_noi - sum(int_monthly) - sum(prin_monthly) - sum(capex)
curr(ws3, r, 14, annual_ncf, font=Font(bold=True, size=12, color="1F3864"), border=THICK_BORDER)

# Alt row shading for Sheet3
for row_idx in range(4, r + 1):
    shade_alt(ws3, row_idx, 14)


# ── Save ─────────────────────────────────────────────────────────────────────
outpath = "output/Roxster_Financial_Analysis_Mar2026.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
