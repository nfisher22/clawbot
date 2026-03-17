
import os
import io
import json
import requests
import msal
import openpyxl
import pandas as pd
from datetime import datetime, timezone
from llm_client import chat_with_fallback
from vault_secrets import get_secrets
get_secrets()

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
MS_USER_EMAIL       = os.getenv("MS_USER_EMAIL", "nfisher@peak10group.com")
OPENAI_API_KEY      = os.getenv("OPENAI_API_KEY")
AUDIT_LOG           = "/opt/clawbot/logs/audit.log"
GRAPH               = "https://graph.microsoft.com/v1.0"

# ---------------------------------------------------------------------------
# Known financial file targets across both sites
# ---------------------------------------------------------------------------
FINANCIAL_FILES = {
    "peak10": [
        ("Quarterly Reports",                              None),
        ("Bookkeeping Property Reports - Balance Sheet project", None),
        ("04.30.25 Fraud Log.xlsx",                        None),
        ("P10 Investment Quarter Report Check List.xlsx",   None),
        ("Vendor Pmt Method 12.17.24.xlsx",                None),
    ],
    "dayton": [
        ("Accounting FW Grafton",                          None),
        ("Fixed Assets",                                   None),
        ("Utility Analysis",                               None),
        ("Community Reports 2024",                         None),
    ],
}

CFO_SYSTEM_PROMPT = """You are the AI CFO, Bookkeeper and Financial Analyst for Peak 10 Group, \
a property management company based in Dayton, Ohio, led by Nathan Fisher.

Your responsibilities:
- Summarise and analyse financial data from SharePoint, email, and the knowledge base
- Generate clear, concise executive summaries and financial reports
- Identify trends, anomalies, and risks across periods and properties
- Compare periods (month-over-month, quarter-over-quarter, year-over-year)
- Flag delinquencies, unusual expenses, or budget variances
- Present numbers in formatted tables where possible

Properties in portfolio include residential communities in the Dayton, OH area \
(FW Grafton, Executive House, Park Layne, Rockwood, Riverstone, 617 Riverview, etc.).

Always be analytical, precise, and proactive. If data is incomplete, say so and \
explain what additional information would improve the analysis.
Today's date: """ + datetime.now().strftime("%B %d, %Y")

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _audit(level, msg):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    try:
        with open(AUDIT_LOG, "a") as f:
            f.write(f"{ts} | {level} | cfo-agent | {msg}\n")
    except Exception:
        pass

def _get_token():
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{AZURE_TENANT_ID}",
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return result.get("access_token")

def _headers():
    token = _get_token()
    if not token:
        raise RuntimeError("CFO agent: could not acquire Graph token.")
    return {"Authorization": "Bearer " + token}

def _resolve_site(alias):
    """Thin wrapper — reuse sharepoint_agent resolver."""
    from sharepoint_agent import _resolve_site as sp_resolve, _headers as sp_headers
    return sp_resolve(alias, sp_headers())

def _download_bytes(site_id, file_path):
    """Download a file from SharePoint and return raw bytes."""
    h = _headers()
    url = f"{GRAPH}/sites/{site_id}/drive/root:/{file_path}:/content"
    resp = requests.get(url, headers=h, allow_redirects=True)
    if resp.status_code != 200:
        return None, f"HTTP {resp.status_code}"
    return resp.content, None

def _xlsx_to_text(raw_bytes, max_rows=200):
    """Parse an xlsx file and return a compact text representation."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    rows.append(f"... (truncated at {max_rows} rows)")
                    break
                if any(c is not None for c in row):
                    rows.append("\t".join("" if c is None else str(c) for c in row))
            if rows:
                parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except Exception as e:
        return f"(Could not parse xlsx: {e})"

def _get_financial_emails(limit=10):
    """Fetch recent emails with financial keywords."""
    try:
        h = _headers()
        filter_q = "contains(subject,'financial') or contains(subject,'invoice') or " \
                   "contains(subject,'payment') or contains(subject,'budget') or " \
                   "contains(subject,'report') or contains(subject,'delinquency') or " \
                   "contains(subject,'rent') or contains(subject,'expense')"
        url = (f"{GRAPH}/users/{MS_USER_EMAIL}/messages"
               f"?$filter={requests.utils.quote(filter_q)}"
               f"&$top={limit}&$select=subject,from,receivedDateTime,bodyPreview"
               f"&$orderby=receivedDateTime desc")
        items = requests.get(url, headers=h).json().get("value", [])
        if not items:
            return ""
        lines = ["=== Recent Financial Emails ==="]
        for e in items:
            lines.append(
                f"FROM: {e.get('from',{}).get('emailAddress',{}).get('address','')}\n"
                f"SUBJECT: {e.get('subject','')}\n"
                f"DATE: {e.get('receivedDateTime','')[:10]}\n"
                f"PREVIEW: {e.get('bodyPreview','')[:300]}\n---"
            )
        return "\n".join(lines)
    except Exception as e:
        return f"(Email fetch error: {e})"

def _get_brain_context(query):
    """Pull relevant vectors from ChromaDB."""
    try:
        import chromadb
        from chromadb.config import Settings
        from openai import OpenAI
        oai = OpenAI(api_key=OPENAI_API_KEY)
        chroma = chromadb.PersistentClient(
            path="/opt/clawbot/chroma_db",
            settings=Settings(anonymized_telemetry=False)
        )
        col = chroma.get_or_create_collection("nate_brain")
        emb = oai.embeddings.create(
            model=os.getenv("EMBED_MODEL", "text-embedding-3-small"),
            input=query
        ).data[0].embedding
        results = col.query(query_embeddings=[emb], n_results=3,
                            include=["documents", "metadatas"])
        docs = results["documents"][0]
        metas = results["metadatas"][0]
        if not docs:
            return ""
        lines = ["=== Knowledge Base ==="]
        for doc, meta in zip(docs, metas):
            lines.append(f"[{meta.get('source','unknown')}]\n{doc[:500]}\n---")
        return "\n".join(lines)
    except Exception as e:
        return f"(Brain error: {e})"

def _fetch_key_spreadsheets(site_alias, paths, max_files=3):
    """Download and parse up to max_files xlsx files from a site."""
    try:
        sid = _resolve_site(site_alias)
        if not sid:
            return f"(Could not resolve site: {site_alias})"
        h = _headers()
        parts = []
        count = 0
        for (name, _) in paths:
            if count >= max_files:
                break
            if not name.endswith(".xlsx"):
                continue
            raw, err = _download_bytes(sid, name)
            if err:
                parts.append(f"[{name}: {err}]")
                continue
            text = _xlsx_to_text(raw)
            parts.append(f"=== {name} ===\n{text[:3000]}")
            count += 1
        return "\n\n".join(parts) if parts else "(No spreadsheets loaded)"
    except Exception as e:
        return f"(Spreadsheet fetch error: {e})"

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def cfo_query(query, include_email=True, include_brain=True):
    """
    Main entry point. Gather financial context from SharePoint, email,
    and the knowledge base, then have the CFO LLM analyse it.
    """
    _audit("INFO", f"cfo_query: {query[:80]}")
    context_parts = []

    # SharePoint spreadsheets from both sites
    xlsx_peak10 = _fetch_key_spreadsheets("peak10", FINANCIAL_FILES["peak10"])
    xlsx_dayton = _fetch_key_spreadsheets("dayton", FINANCIAL_FILES["dayton"])
    if xlsx_peak10 and xlsx_peak10 != "(No spreadsheets loaded)":
        context_parts.append("=== Peak 10 Main — Financial Files ===\n" + xlsx_peak10)
    if xlsx_dayton and xlsx_dayton != "(No spreadsheets loaded)":
        context_parts.append("=== P10 Mgmt Dayton — Financial Files ===\n" + xlsx_dayton)

    if include_email:
        email_ctx = _get_financial_emails()
        if email_ctx:
            context_parts.append(email_ctx)

    if include_brain:
        brain_ctx = _get_brain_context(query)
        if brain_ctx:
            context_parts.append(brain_ctx)

    full_context = "\n\n".join(context_parts) if context_parts else "(No financial data loaded)"

    messages = [
        {"role": "system", "content": CFO_SYSTEM_PROMPT + "\n\nFinancial data:\n\n" + full_context},
        {"role": "user", "content": query},
    ]

    result = chat_with_fallback(messages)
    _audit("SUCCESS", f"cfo_query completed")
    return result


def cfo_summary():
    """Generate a high-level financial executive summary."""
    return cfo_query(
        "Generate a concise executive financial summary for Peak 10 Group. "
        "Include: key financial figures, any trends or anomalies, delinquencies, "
        "and top action items. Format with clear sections and bullet points."
    )


def cfo_report(topic):
    """Generate a focused financial report on a specific topic."""
    return cfo_query(
        f"Generate a detailed financial report on the following topic for Peak 10 Group: {topic}. "
        "Include relevant data, trends, comparisons where possible, and recommendations."
    )


def cfo_read_file(site_alias, file_path, question=None):
    """Download a specific financial file and analyse it."""
    _audit("INFO", f"cfo_read_file: {site_alias}/{file_path}")
    sid = _resolve_site(site_alias)
    if not sid:
        return f"Could not find site: {site_alias}"

    raw, err = _download_bytes(sid, file_path)
    if err:
        return f"Could not download {file_path}: {err}"

    if file_path.endswith(".xlsx"):
        content = _xlsx_to_text(raw)
    else:
        content = raw.decode("utf-8", errors="replace")[:5000]

    prompt = question or f"Analyse this financial document and provide a detailed summary with key figures, trends, and any anomalies."
    messages = [
        {"role": "system", "content": CFO_SYSTEM_PROMPT},
        {"role": "user", "content": f"File: {file_path}\n\n{content}\n\n{prompt}"},
    ]
    return chat_with_fallback(messages)



def cfo_folder(site_alias, folder_path, question=None):
    """
    Pull all xlsx/pdf/csv files from a specific SharePoint folder
    and run CFO analysis on them.

    Args:
        site_alias:  'peak10', 'dayton', or any alias/site name
        folder_path: folder path within the drive root
        question:    optional specific question; defaults to general analysis
    """
    _audit("INFO", f"cfo_folder: {site_alias}/{folder_path}")
    sid = _resolve_site(site_alias)
    if not sid:
        return f"Could not find site: {site_alias}"

    h = _headers()
    # List files in the folder
    url = f"{GRAPH}/sites/{sid}/drive/root:/{folder_path}:/children?$top=50&$select=name,size,file,webUrl"
    resp = requests.get(url, headers=h).json()
    if "error" in resp:
        return f"Could not list folder '{folder_path}': {resp['error'].get('message', resp['error'])}"

    items = [i for i in resp.get("value", []) if "file" in i]
    if not items:
        return f"No files found in {site_alias}/{folder_path}."

    SUPPORTED = (".xlsx", ".xls", ".csv", ".pdf")
    parseable = [i for i in items if i["name"].lower().endswith(SUPPORTED)]
    skipped   = [i for i in items if not i["name"].lower().endswith(SUPPORTED)]

    if not parseable:
        names = ", ".join(i["name"] for i in items[:10])
        return f"Found {len(items)} file(s) in {folder_path} but none are xlsx/csv/pdf. Files: {names}"

    context_parts = [f"Folder: {site_alias}/{folder_path}\n"]

    for item in parseable[:5]:   # cap at 5 files to stay within token limits
        fname = item["name"]
        raw, err = _download_bytes(sid, f"{folder_path}/{fname}")
        if err:
            context_parts.append(f"[{fname}: download failed — {err}]")
            continue
        ext = fname.lower().rsplit(".", 1)[-1]
        if ext in ("xlsx", "xls"):
            text = _xlsx_to_text(raw)[:3000]
        elif ext == "csv":
            text = raw.decode("utf-8", errors="replace")[:3000]
        elif ext == "pdf":
            try:
                import pdfplumber, io as _io
                with pdfplumber.open(_io.BytesIO(raw)) as pdf:
                    pages = [p.extract_text() or "" for p in pdf.pages[:20]]
                text = "\n".join(pages)[:3000]
            except Exception as e:
                text = f"(PDF parse error: {e})"
        else:
            text = raw.decode("utf-8", errors="replace")[:3000]
        context_parts.append(f"=== {fname} ===\n{text}")

    if skipped:
        context_parts.append(f"\n(Skipped {len(skipped)} non-financial file(s): "
                             + ", ".join(i["name"] for i in skipped[:5]) + ")")
    if len(parseable) > 5:
        context_parts.append(f"\n(Note: Only analysed 5 of {len(parseable)} files. "
                             "Use /cfo read to target a specific file.)")

    full_context = "\n\n".join(context_parts)
    prompt = question or (
        f"Analyse all files in the folder '{folder_path}' for Peak 10 Group. "
        "Summarise key financial figures, identify trends, flag anomalies, "
        "and provide actionable insights."
    )

    messages = [
        {"role": "system", "content": CFO_SYSTEM_PROMPT + "\n\nFolder contents:\n\n" + full_context},
        {"role": "user", "content": prompt},
    ]
    result = chat_with_fallback(messages)
    _audit("SUCCESS", f"cfo_folder completed: {site_alias}/{folder_path}")
    return result


if __name__ == "__main__":
    print("=== CFO Agent Smoke Test ===")
    print(cfo_summary())
