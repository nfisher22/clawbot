import os
import json
import logging
import requests
from pathlib import Path
from datetime import datetime, timezone
from dotenv import load_dotenv
from openai import OpenAI

import chromadb
from chromadb.config import Settings
from tavily import TavilyClient
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
import telegram.error
import msal
from llm_client import chat_with_fallback
from agents import route, email_agent, calendar_agent, research_agent, brain_agent
from onedrive_agent import list_files, save_file, search_files, save_daily_summary
from sharepoint_agent import list_sites, list_folders, search_files as sp_search_files, read_file as sp_read_file, save_file as sp_save_file
from cfo_agent import cfo_query, cfo_summary, cfo_report, cfo_read_file, cfo_folder
from teams_agent import get_my_chats, get_chat_messages, get_teams_list

from vault_secrets import get_secrets
from task_checkpoint import add_task, complete_task, get_pending_tasks, update_task_status, get_task_summary
get_secrets()
load_dotenv()  # fallback

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
MODEL_NAME = os.getenv("MODEL_NAME", "gpt-4o-mini")
EMBED_MODEL = os.getenv("EMBED_MODEL", "text-embedding-3-small")
TAVILY_API_KEY = os.getenv("TAVILY_API_KEY")
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN") or os.getenv("BOT_TOKEN")
AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
MS_USER_EMAIL = os.getenv("MS_USER_EMAIL", "nfisher@peak10group.com")

DATA_DIR = Path("/opt/clawbot/data")
MEMORY_DIR = Path("/opt/clawbot/memory")
CHROMA_DIR = "/opt/clawbot/chroma_db"
AUDIT_LOG = "/opt/clawbot/logs/audit.log"
SESSION_FILE = "/opt/clawbot/logs/session_state.json"

client = OpenAI(api_key=OPENAI_API_KEY)
tavily = TavilyClient(api_key=TAVILY_API_KEY)

chroma_client = chromadb.PersistentClient(path=CHROMA_DIR, settings=Settings(anonymized_telemetry=False))
collection = chroma_client.get_or_create_collection(name="nate_brain")

logging.basicConfig(level=logging.INFO)
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("telegram").setLevel(logging.WARNING)
conversation_history = {}

def audit(level, script, message):
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    line = f"{ts} | {level} | {script} | {message}\n"
    try:
        with open(AUDIT_LOG, "a") as f:
            f.write(line)
    except Exception:
        pass

def save_session():
    try:
        with open(SESSION_FILE, "w") as f:
            json.dump(conversation_history, f)
    except Exception as e:
        audit("ERROR", "telegram-bot", f"Failed to save session: {e}")

def load_session():
    try:
        if os.path.exists(SESSION_FILE):
            with open(SESSION_FILE, "r") as f:
                data = json.load(f)
                # Convert string keys back to int
                return {int(k): v for k, v in data.items()}
    except Exception as e:
        audit("ERROR", "telegram-bot", f"Failed to load session: {e}")
    return {}

def get_graph_token():
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority="https://login.microsoftonline.com/" + AZURE_TENANT_ID,
        client_credential=AZURE_CLIENT_SECRET
    )
    result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    return result.get("access_token")

def get_recent_emails(count=5):
    token = get_graph_token()
    if not token:
        return "Could not authenticate with Microsoft."
    headers = {"Authorization": "Bearer " + token}
    url = "https://graph.microsoft.com/v1.0/users/" + MS_USER_EMAIL + "/messages?$top=" + str(count) + "&$select=subject,from,receivedDateTime,bodyPreview"
    response = requests.get(url, headers=headers).json()
    emails = response.get("value", [])
    if not emails:
        return "No emails found."
    output = []
    for e in emails:
        output.append("FROM: " + e["from"]["emailAddress"]["address"] + "\nSUBJECT: " + e["subject"] + "\nDATE: " + e["receivedDateTime"] + "\nPREVIEW: " + e["bodyPreview"][:200])
    return "\n---\n".join(output)

def get_calendar_events(count=10):
    token = get_graph_token()
    if not token:
        return "Could not authenticate with Microsoft."
    headers = {"Authorization": "Bearer " + token}
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    url = ("https://graph.microsoft.com/v1.0/users/" + MS_USER_EMAIL +
           "/calendarView?startDateTime=" + now +
           "&endDateTime=2026-03-15T00:00:00Z" +
           "&$select=subject,start,end,location,organizer&$top=" + str(count) +
           "&$orderby=start/dateTime")
    response = requests.get(url, headers=headers).json()
    events = response.get("value", [])
    if not events:
        return "No upcoming calendar events found."
    output = []
    for e in events:
        start = e.get("start", {}).get("dateTime", "")[:16].replace("T", " ")
        end = e.get("end", {}).get("dateTime", "")[:16].replace("T", " ")
        subject = e.get("subject", "No title")
        location = e.get("location", {}).get("displayName", "")
        organizer = e.get("organizer", {}).get("emailAddress", {}).get("name", "")
        line = "MEETING: " + subject + "\nSTART: " + start + "\nEND: " + end
        if location:
            line += "\nLOCATION: " + location
        if organizer:
            line += "\nORGANIZER: " + organizer
        output.append(line)
    return "\n---\n".join(output)

def create_calendar_event(subject, start_dt, end_dt, location="", body=""):
    token = get_graph_token()
    if not token:
        return "Could not authenticate with Microsoft."
    headers = {"Authorization": "Bearer " + token, "Content-Type": "application/json"}
    payload = {
        "subject": subject,
        "start": {"dateTime": start_dt, "timeZone": "America/New_York"},
        "end": {"dateTime": end_dt, "timeZone": "America/New_York"},
        "location": {"displayName": location},
        "body": {"contentType": "Text", "content": body}
    }
    url = "https://graph.microsoft.com/v1.0/users/" + MS_USER_EMAIL + "/events"
    response = requests.post(url, headers=headers, json=payload)
    if response.status_code in [200, 201]:
        return "Meeting created successfully."
    return "Failed to create meeting: " + str(response.json())

def load_text_file(path):
    if not path.exists():
        return ""
    try:
        return path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return ""

def load_memory():
    long_term = load_text_file(MEMORY_DIR / "MEMORY.md")
    today = datetime.now().strftime("%Y-%m-%d")
    daily = load_text_file(MEMORY_DIR / f"{today}.md")
    return long_term, daily

def write_daily_memory(entry):
    today = datetime.now().strftime("%Y-%m-%d")
    daily_path = MEMORY_DIR / f"{today}.md"
    ts = datetime.now().strftime("%H:%M")
    try:
        with open(daily_path, "a") as f:
            f.write(f"\n## {ts}\n{entry}\n")
    except Exception as e:
        audit("ERROR", "telegram-bot", f"Failed to write daily memory: {e}")

def load_system_context():
    system_prompt = load_text_file(DATA_DIR / "SYSTEM_PROMPT.md")
    profile = load_text_file(DATA_DIR / "PROFILE.md")
    ea_workflow = load_text_file(DATA_DIR / "EA_WORKFLOW.md")
    long_term, daily = load_memory()
    sep = "\n\n"
    return (system_prompt + sep +
            "User Profile:" + sep + profile + sep +
            "EA Workflow:" + sep + ea_workflow + sep +
            "Long-Term Memory:" + sep + long_term + sep +
            "Today's Log:" + sep + daily).strip()

def get_embedding(text):
    response = client.embeddings.create(model=EMBED_MODEL, input=text)
    return response.data[0].embedding

def search_brain(query, n_results=5):
    embedding = get_embedding(query)
    results = collection.query(query_embeddings=[embedding], n_results=n_results, include=["documents", "metadatas"])
    chunks = []
    for doc, meta in zip(results["documents"][0], results["metadatas"][0]):
        source = meta.get("source", "unknown")
        chunks.append("FILE: " + source + "\nCONTENT:\n" + doc + "\n")
    return "\n---\n".join(chunks)

def needs_web_search(query):
    keywords = ["today", "current", "latest", "news", "price", "weather", "right now", "this week", "stock", "score", "who won", "near me", "nearby", "open now", "movies", "showtimes", "restaurant", "search for", "find me", "look up"]
    return any(k in query.lower() for k in keywords)

def needs_email(query):
    keywords = ["email", "inbox", "mail", "message from", "did i get", "unread", "emails"]
    return any(k in query.lower() for k in keywords)

def needs_calendar(query):
    keywords = ["calendar", "schedule", "meeting", "appointment", "what do i have", "upcoming", "agenda", "when is", "add meeting", "create meeting", "schedule meeting"]
    return any(k in query.lower() for k in keywords)

def needs_create_event(query):
    keywords = ["add meeting", "create meeting", "schedule meeting", "add to calendar", "create event", "book meeting"]
    return any(k in query.lower() for k in keywords)

def needs_memory_write(query):
    keywords = ["remember this", "remember that", "don't forget", "make a note", "note that", "log this"]
    return any(k in query.lower() for k in keywords)

def search_web(query):
    results = tavily.search(query)
    output = []
    for r in results.get("results", [])[:3]:
        output.append("SOURCE: " + r["url"] + "\n" + r["content"] + "\n")
    return "\n---\n".join(output)

def chat(user_id, user_input):
    if user_id not in conversation_history:
        conversation_history[user_id] = []
    history = conversation_history[user_id]
    system_context = load_system_context()

    if needs_memory_write(user_input):
        write_daily_memory(f"User asked to remember: {user_input}")
        audit("INFO", "telegram-bot", f"Memory write triggered: {user_input[:50]}")

    if needs_email(user_input):
        context = "RECENT EMAILS:\n" + get_recent_emails()
        audit("INFO", "telegram-bot", "Email context loaded")
    elif needs_create_event(user_input):
        context = "CALENDAR ACTION: User wants to create a meeting. Ask for details if not provided: subject, date, start time, end time, location."
    elif needs_calendar(user_input):
        context = "UPCOMING CALENDAR EVENTS:\n" + get_calendar_events()
        audit("INFO", "telegram-bot", "Calendar context loaded")
    elif needs_web_search(user_input):
        context = "WEB SEARCH RESULTS:\n" + search_web(user_input + " Columbus Ohio")
        audit("INFO", "telegram-bot", f"Web search: {user_input[:50]}")
    else:
        context = "RELEVANT BRAIN CONTEXT:\n" + search_brain(user_input)
        audit("INFO", "telegram-bot", f"Brain search: {user_input[:50]}")

    system_message = system_context + "\n\n" + context
    history.append({"role": "user", "content": user_input})

    # Keep history to last 20 messages to avoid context overflow
    if len(history) > 20:
        history = history[-20:]
        conversation_history[user_id] = history

    output = chat_with_fallback([{"role": "system", "content": system_message}] + history)
    history.append({"role": "assistant", "content": output})

    # Save session after every message
    save_session()

    if len(output) > 100 and not needs_web_search(user_input):
        write_daily_memory(f"Q: {user_input[:100]}\nA: {output[:200]}")

    return output

async def cmd_email(update, context):
    query = " ".join(context.args) if context.args else "Show my latest emails"
    await update.message.reply_text("Checking emails...")
    result = email_agent(query)
    await update.message.reply_text(result)

async def cmd_calendar(update, context):
    query = " ".join(context.args) if context.args else "What is on my calendar this week"
    await update.message.reply_text("Checking calendar...")
    result = calendar_agent(query)
    await update.message.reply_text(result)

async def cmd_research(update, context):
    query = " ".join(context.args) if context.args else "Latest news Columbus Ohio"
    await update.message.reply_text("Researching...")
    result = research_agent(query)
    await update.message.reply_text(result)

async def cmd_brain(update, context):
    query = " ".join(context.args) if context.args else "What do you know about me"
    await update.message.reply_text("Searching knowledge base...")
    result = brain_agent(query)
    await update.message.reply_text(result)

async def cmd_onedrive(update, context):
    args = context.args
    if not args:
        result = list_files()
        await update.message.reply_text("📁 OneDrive Root:\n" + result)
        return
    subcmd = args[0].lower()
    if subcmd == "list":
        folder = " ".join(args[1:]) if len(args) > 1 else "root"
        result = list_files(folder)
        await update.message.reply_text(f"📁 OneDrive {folder}:\n" + result)
    elif subcmd == "search":
        query = " ".join(args[1:])
        result = search_files(query)
        await update.message.reply_text("🔍 Search results:\n" + result)
    elif subcmd == "save":
        content_to_save = " ".join(args[1:])
        ts = __import__("datetime").datetime.now().strftime("%Y-%m-%d-%H%M")
        result = save_file(f"note-{ts}.txt", content_to_save)
        await update.message.reply_text(result)
    elif subcmd == "sync":
        result = save_daily_summary()
        await update.message.reply_text(result)
    else:
        await update.message.reply_text("Usage: /onedrive list [folder] | search [query] | save [note] | sync")

async def cmd_cfo(update, context):
    args = context.args
    if not args:
        await update.message.reply_text("CFO agent thinking... this may take 15-20 seconds.")
        result = cfo_summary()
        await update.message.reply_text(result[:4096])
        return
    subcmd = args[0].lower()
    if subcmd == "summary":
        await update.message.reply_text("Generating executive summary...")
        result = cfo_summary()
        await update.message.reply_text(result[:4096])
    elif subcmd == "report":
        topic = " ".join(args[1:]) if len(args) > 1 else "overall financial performance"
        await update.message.reply_text(f"Generating report on: {topic}...")
        result = cfo_report(topic)
        await update.message.reply_text(result[:4096])
    elif subcmd == "read":
        # /cfo read [site] [file path]
        if len(args) < 3:
            await update.message.reply_text("Usage: /cfo read [site] [file/path.xlsx]")
            return
        site = args[1]
        file_path = " ".join(args[2:])
        await update.message.reply_text(f"Reading {file_path}...")
        result = cfo_read_file(site, file_path)
        await update.message.reply_text(result[:4096])
    elif subcmd == "folder":
        # /cfo folder [site] [folder path] -- optional :: question
        # e.g. /cfo folder peak10 Quarterly Reports
        # e.g. /cfo folder dayton Accounting FW Grafton::What are the biggest expenses?
        if len(args) < 3:
            await update.message.reply_text(
                "Usage: /cfo folder [site] [folder path]\n"
                "Optionally append a question: /cfo folder dayton Accounting FW Grafton::What are total expenses?"
            )
            return
        raw = " ".join(args[2:])
        if "::" in raw:
            folder_path, question = raw.split("::", 1)
            folder_path = folder_path.strip()
            question = question.strip()
        else:
            folder_path = raw.strip()
            question = None
        site = args[1]
        await update.message.reply_text(f"Analysing folder: {site}/{folder_path}...")
        result = cfo_folder(site, folder_path, question)
        await update.message.reply_text(result[:4096])
    else:
        # Treat everything as a natural language financial question
        question = " ".join(args)
        await update.message.reply_text("CFO agent analysing...")
        result = cfo_query(question)
        await update.message.reply_text(result[:4096])


async def cmd_sharepoint(update, context):
    args = context.args
    if not args:
        result = list_sites()
        await update.message.reply_text("SharePoint Sites:\n" + result)
        return
    subcmd = args[0].lower()
    if subcmd == "sites":
        result = list_sites()
        await update.message.reply_text("SharePoint Sites:\n" + result)
    elif subcmd == "list":
        # Join everything after "list" as site name; use "::" to separate site::folder
        # e.g. /sharepoint list P10 Mgmt Dayton
        # e.g. /sharepoint list P10 Mgmt Dayton::General/Reports
        raw = " ".join(args[1:]) if len(args) > 1 else "Peak10"
        if "::" in raw:
            site, folder = raw.split("::", 1)
            site = site.strip()
            folder = folder.strip() or None
        else:
            site = raw.strip()
            folder = None
        result = list_folders(site, folder)
        label = f"{site}/{folder}" if folder else site
        await update.message.reply_text(f"SharePoint - {label}:\n" + result)
    elif subcmd == "search":
        if len(args) < 2:
            await update.message.reply_text("Usage: /sharepoint search [query]")
            return
        result = sp_search_files(" ".join(args[1:]))
        await update.message.reply_text("SharePoint Search:\n" + result)
    elif subcmd == "read":
        if len(args) < 3:
            await update.message.reply_text("Usage: /sharepoint read [site] [file/path]")
            return
        site = args[1]
        file_path = " ".join(args[2:])
        result = sp_read_file(site, file_path)
        await update.message.reply_text(f"{file_path}:\n" + result[:4000])
    else:
        await update.message.reply_text(
            "SharePoint commands:\n"
            "  /sharepoint - list all sites\n"
            "  /sharepoint sites - list all sites\n"
            "  /sharepoint list [site] [folder] - browse files\n"
            "  /sharepoint search [query] - search across SharePoint\n"
            "  /sharepoint read [site] [path] - read a file"
        )


async def cmd_teams(update, context):
    args = context.args
    if not args:
        result = get_my_chats()
        await update.message.reply_text("💬 Your Teams Chats:\n" + result)
        return
    subcmd = args[0].lower()
    if subcmd == "chats":
        result = get_my_chats()
        await update.message.reply_text("💬 Your Teams Chats:\n" + result)
    elif subcmd == "teams":
        result = get_teams_list()
        await update.message.reply_text("👥 Your Teams:\n" + result)
    elif subcmd == "messages":
        if len(args) < 2:
            await update.message.reply_text("Usage: /teams messages [chat_id]")
            return
        chat_id = args[1]
        result = get_chat_messages(chat_id)
        await update.message.reply_text("💬 Messages:\n" + result)
    else:
        await update.message.reply_text("Usage: /teams chats | teams | messages [chat_id]")

async def start(update, context):
    audit("INFO", "telegram-bot", f"New session started by user {update.effective_user.id}")
    await update.message.reply_text("ClawBot is live. Ask me anything.")

async def clear(update, context):
    user_id = update.effective_user.id
    conversation_history[user_id] = []
    save_session()
    audit("INFO", "telegram-bot", f"Conversation cleared by user {user_id}")
    await update.message.reply_text("Conversation history cleared.")

async def handle_document(update, context):
    """Handle file attachments -- route xlsx/pdf/csv to CFO agent."""
    doc = update.message.document
    if not doc:
        return
    fname = doc.file_name or ""
    ext = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext not in ("xlsx", "xls", "csv", "pdf"):
        await update.message.reply_text(
            f"I received '{fname}' but I can only analyse xlsx, csv, or pdf files for now."
        )
        return

    await update.message.reply_text(f"Received {fname} -- analysing with CFO agent...")

    # Download from Telegram
    tg_file = await context.bot.get_file(doc.file_id)
    import io, requests as req
    resp = req.get(tg_file.file_path)
    raw_bytes = resp.content

    # Parse based on type
    if ext in ("xlsx", "xls"):
        import openpyxl, io as _io
        try:
            wb = openpyxl.load_workbook(_io.BytesIO(raw_bytes), data_only=True)
            parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 300: rows.append("...(truncated)"); break
                    if any(c is not None for c in row):
                        rows.append("\t".join("" if c is None else str(c) for c in row))
                if rows:
                    parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
            file_text = "\n\n".join(parts)[:6000]
        except Exception as e:
            file_text = f"(Could not parse xlsx: {e})"
    elif ext == "csv":
        file_text = raw_bytes.decode("utf-8", errors="replace")[:6000]
    elif ext == "pdf":
        try:
            import pdfplumber, io as _io
            with pdfplumber.open(_io.BytesIO(raw_bytes)) as pdf:
                pages = [p.extract_text() or "" for p in pdf.pages[:20]]
            file_text = "\n".join(pages)[:6000]
        except Exception as e:
            file_text = f"(Could not parse pdf: {e})"
    else:
        file_text = raw_bytes.decode("utf-8", errors="replace")[:6000]

    # Get optional question from caption
    question = update.message.caption or "Analyse this financial document and provide a detailed summary with key figures, trends, and any anomalies."

    from llm_client import chat_with_fallback
    from cfo_agent import CFO_SYSTEM_PROMPT
    messages = [
        {"role": "system", "content": CFO_SYSTEM_PROMPT},
        {"role": "user", "content": f"File: {fname}\n\n{file_text}\n\n{question}"},
    ]
    result = chat_with_fallback(messages)
    await update.message.reply_text(result[:4096])


async def handle_message(update, context):
    with open("/opt/clawbot/logs/chatid.log", "a") as f:
        f.write(str(update.effective_chat.id) + " | " + str(update.effective_user.first_name) + "\n")
    user_id = update.effective_user.id
    user_input = update.message.text
    await update.message.reply_text("Thinking...")
    try:
        output = chat(user_id, user_input)
        await update.message.reply_text(output)
    except Exception as e:
        audit("ERROR", "telegram-bot", f"Exception: {str(e)}")
        await update.message.reply_text("Error: " + str(e))




async def cmd_meetings(update, context):
    user_id = update.effective_user.id
    if user_id != 8647502718:
        await update.message.reply_text("Unauthorized.")
        return
    args = context.args
    source = args[0].lower() if args else "all"
    try:
        lines = []
        if source in ["all", "fathom"]:
            from fathom_agent import get_recent_meetings, format_meetings
            fathom_meetings = get_recent_meetings(3)
            if fathom_meetings:
                lines.append(format_meetings(fathom_meetings))
        if source in ["all", "fireflies"]:
            from fireflies_agent import get_recent_transcripts, format_transcripts
            ff_transcripts = get_recent_transcripts(3)
            if ff_transcripts:
                lines.append(format_transcripts(ff_transcripts))
        if lines:
            await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        else:
            await update.message.reply_text("No recent meetings found.")
    except Exception as e:
        await update.message.reply_text(f"Error fetching meetings: {e}")

async def cmd_task(update, context):
    user_id = update.effective_user.id
    if user_id != 8647502718:
        await update.message.reply_text("Unauthorized.")
        return
    args = context.args
    if not args:
        summary = get_task_summary()
        await update.message.reply_text(summary, parse_mode="Markdown")
        return
    if args[0] == "done" and len(args) > 1:
        try:
            task_id = int(args[1])
            success = complete_task(task_id)
            if success:
                await update.message.reply_text(f"✅ Task #{task_id} marked complete.")
            else:
                await update.message.reply_text(f"Task #{task_id} not found.")
        except ValueError:
            await update.message.reply_text("Usage: /task done <id>")
        return
    if args[0] == "high" and len(args) > 1:
        title = " ".join(args[1:])
        task_id = add_task(title, title, source="telegram", priority="high")
        await update.message.reply_text(f"🔴 High priority task #{task_id} added: {title}")
        return
    if args[0] == "low" and len(args) > 1:
        title = " ".join(args[1:])
        task_id = add_task(title, title, source="telegram", priority="low")
        await update.message.reply_text(f"⚪ Low priority task #{task_id} added: {title}")
        return
    if args[0] == "status" and len(args) > 2:
        try:
            task_id = int(args[1])
            status = args[2]
            success = update_task_status(task_id, status)
            if success:
                await update.message.reply_text(f"✅ Task #{task_id} → {status}")
            else:
                await update.message.reply_text(f"Task #{task_id} not found.")
        except ValueError:
            await update.message.reply_text("Usage: /task status <id> <status>")
        return
    # Add normal priority task
    title = " ".join(args)
    task_id = add_task(title, title, source="telegram", priority="normal")
    await update.message.reply_text(f"🟡 Task #{task_id} added: {title}")

CORRECTIONS_LOG = "/opt/clawbot/logs/corrections.log"

async def cmd_correct(update, context):
    user_id = update.effective_user.id
    if user_id != 8647502718:
        await update.message.reply_text("Unauthorized.")
        return
    correction_text = " ".join(context.args) if context.args else ""
    if not correction_text:
        await update.message.reply_text("Usage: /correct <describe what was wrong>")
        return
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    # Get last bot message for context
    history = conversation_history.get(user_id, [])
    last_exchange = ""
    if len(history) >= 2:
        last_exchange = f"User: {history[-2].get('content','')}\nBot: {history[-1].get('content','')}"
    entry = f"{ts} | CORRECTION | {correction_text}\nCONTEXT: {last_exchange[:500]}\n---\n"
    try:
        with open(CORRECTIONS_LOG, "a") as f:
            f.write(entry)
        audit("INFO", "telegram-bot", f"Correction logged: {correction_text[:100]}")
        await update.message.reply_text("✅ Correction logged. Thanks — I'll use this to improve.")
    except Exception as e:
        await update.message.reply_text(f"Failed to log correction: {e}")


def main():
    global conversation_history
    conversation_history = load_session()
    try:
        from main import quick_load_vectors
        quick_load_vectors()
    except Exception as e:
        print(f'Quick loader skipped: {e}')
    audit("INFO", "telegram-bot", f"ClawBot V1.5 started — session restored with {len(conversation_history)} users")
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("clear", clear))
    app.add_handler(CommandHandler("email", cmd_email))
    app.add_handler(CommandHandler("calendar", cmd_calendar))
    app.add_handler(CommandHandler("research", cmd_research))
    app.add_handler(CommandHandler("brain", cmd_brain))
    app.add_handler(CommandHandler("onedrive", cmd_onedrive))
    app.add_handler(CommandHandler("cfo", cmd_cfo))
    app.add_handler(CommandHandler("sharepoint", cmd_sharepoint))
    app.add_handler(CommandHandler("teams", cmd_teams))
    app.add_handler(CommandHandler("correct", cmd_correct))
    app.add_handler(CommandHandler("task", cmd_task))
    app.add_handler(CommandHandler("meetings", cmd_meetings))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    print("ClawBot V1.5 Session Recovery + Memory + Brain + Web + Email + Calendar is live...")
    app.run_polling()

if __name__ == "__main__":
    import time
    for _attempt in range(5):
        try:
            main()
            break
        except telegram.error.InvalidToken:
            if _attempt >= 4:
                logging.error("Telegram token invalid after 5 attempts — exiting")
                raise
            _wait = 2 ** _attempt
            logging.warning(f"InvalidToken on attempt {_attempt + 1}/5, retrying in {_wait}s")
            time.sleep(_wait)
            get_secrets()  # re-read from Vault in case token was rotated
            TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN") or os.getenv("BOT_TOKEN")
